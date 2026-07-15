import json
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT_DIR = '/tmp/zh/work/VisionSelector/qwen-evaluation/result/eval_time'
MODEL_NAME = 'VisionSelector-Qwen2.5-VL-3B-train-V2-causal-7epoch'
TASKS = ['docvqa_val', 'chartqa', 'textvqa_val', 'ocrbench', 'scienceqa_img', 'ai2d_no_mask', 'mmmu_val', 'mme', 'pope']

# 收集结果
results = {}
for task in TASKS:
    result_path = f'{ROOT_DIR}/{task}/0.2/{MODEL_NAME}_selector_0.2_3b/output_ckpt__{MODEL_NAME}'
    if os.path.exists(result_path):
        json_files = [f for f in os.listdir(result_path) if f.endswith('_results.json')]
        if json_files:
            with open(os.path.join(result_path, json_files[0]), 'r') as f:
                data = json.load(f)
            task_results = data.get('results', {}).get(task, {})
            results[task] = task_results

# 创建 Excel
wb = Workbook()
ws = wb.active
ws.title = MODEL_NAME

# 设置表头样式
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ['Dataset'] + [t.replace('_val', '').replace('_', ' ').title() for t in TASKS]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 数据行 - 指标映射
metric_map = {
    'docvqa_val': 'anls,none',
    'chartqa': 'relaxed_overall,none',
    'textvqa_val': 'accuracy,none',
    'ocrbench': 'final_score,none',
    'scienceqa_img': 'accuracy,none',
    'ai2d_no_mask': 'accuracy,none',
    'mmmu_val': 'accuracy,none',
    'mme': 'score,none',
    'pope': 'score,none'
}

row = 2
cell = ws.cell(row=row, column=1, value=MODEL_NAME.split('-')[-1])
cell.font = header_font
cell.alignment = center_align
cell.border = thin_border

for col, task in enumerate(TASKS, 2):
    if task in results and metric_map.get(task) in results[task]:
        value = results[task][metric_map[task]]
        # 特殊处理：0-1 范围的值乘以 100
        if isinstance(value, (int, float)) and 0 <= value <= 1:
            value = value * 100
        cell = ws.cell(row=row, column=col, value=round(value, 2))
    else:
        cell = ws.cell(row=row, column=col, value='-')
    cell.alignment = center_align
    cell.border = thin_border

# 调整列宽
ws.column_dimensions['A'].width = 25
for col in range(2, len(TASKS) + 2):
    ws.column_dimensions[chr(ord('A') + col - 1)].width = 15

# 保存
output_file = f'/tmp/zh/work/VisionSelector/qwen-evaluation/{MODEL_NAME}_results.xlsx'
wb.save(output_file)
sys.stdout.write(f'Excel 已生成: {output_file}\n')
sys.stdout.flush()

# 打印结果摘要
sys.stdout.write('\n=== 结果汇总 ===\n')
sys.stdout.flush()
for task in TASKS:
    if task in results and metric_map.get(task) in results[task]:
        value = results[task][metric_map[task]]
        if isinstance(value, (int, float)) and 0 <= value <= 1:
            value = value * 100
        sys.stdout.write(f'{task}: {value:.2f}\n')
    else:
        sys.stdout.write(f'{task}: -\n')
    sys.stdout.flush()
