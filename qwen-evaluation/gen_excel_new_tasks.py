import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ROOT_DIR = '/tmp/zh/work/VisionSelector/qwen-evaluation/result/eval_time'

# 两个模型
MODELS = {
    'Layer-Attn-10epoch': {
        'dir_name': 'VisionSelector-Qwen2.5-VL-3B-train-Layer-Attn-10epoch_selector_0.2_3b',
        'sub_dir': 'output_ckpt__VisionSelector-Qwen2.5-VL-3B-train-Layer-Attn-10epoch',
    },
    'Qwen2.5-VL-3B-Instruct (orig)': {
        'dir_name': 'Qwen2.5-VL-3B-Instruct_orig_0.2_new_preprocess',
        'sub_dir': 'pretrained__Qwen2.5-VL-3B-Instruct',
    },
}

TASKS = ['coco2017_cap_val', 'gqa', 'mmbench_en_dev', 'nocaps_val', 'ok_vqa_val2014']

# 每个数据集选取的指标
metric_map = {
    'coco2017_cap_val': ['coco_CIDEr,none', 'coco_Bleu_1,none', 'coco_Bleu_4,none'],
    'gqa': ['exact_match,none'],
    'mmbench_en_dev': ['gpt_eval_score,none'],
    'nocaps_val': ['nocaps_CIDEr,none', 'nocaps_Bleu_1,none', 'nocaps_Bleu_4,none'],
    'ok_vqa_val2014': ['exact_match,none'],
}

# 收集结果
all_results = {}
for model_label, model_info in MODELS.items():
    results = {}
    for task in TASKS:
        result_dir = f'{ROOT_DIR}/{task}/0.2/{model_info["dir_name"]}/{model_info["sub_dir"]}'
        if os.path.exists(result_dir):
            json_files = [f for f in os.listdir(result_dir) if f.endswith('_results.json')]
            if json_files:
                with open(os.path.join(result_dir, json_files[0]), 'r') as f:
                    data = json.load(f)
                task_results = data.get('results', {}).get(task, {})
                results[task] = task_results
    all_results[model_label] = results

def format_value(results, task, metric_key):
    if task not in results or metric_key not in results[task]:
        return None
    value = results[task][metric_key]
    if not isinstance(value, (int, float)):
        return None
    # Bleu 和 0-1 范围的值乘以 100，CIDEr 不乘
    if 'CIDEr' not in metric_key and ('Bleu' in metric_key or 0 <= value <= 1):
        value = value * 100
    return round(value, 2)

def format_cell(results, task, metrics):
    values = []
    for m in metrics:
        v = format_value(results, task, m)
        if v is not None:
            values.append(f'{v:.2f}')
    if not values:
        return '-'
    return ' / '.join(values)

# 创建 Excel
wb = Workbook()
ws = wb.active
ws.title = 'Results'

# 设置表头样式
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ['Model'] + [t.replace('_val', '').replace('_val2014', '').replace('_', ' ').title() for t in TASKS]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 数据行
for row_idx, (model_label, results) in enumerate(all_results.items(), 2):
    cell = ws.cell(row=row_idx, column=1, value=model_label)
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

    for col, task in enumerate(TASKS, 2):
        metrics = metric_map.get(task, [])
        cell_value = format_cell(results, task, metrics)
        cell = ws.cell(row=row_idx, column=col, value=cell_value)
        cell.alignment = center_align
        cell.border = thin_border

# 调整列宽
ws.column_dimensions['A'].width = 35
for col in range(2, len(TASKS) + 2):
    ws.column_dimensions[chr(ord('A') + col - 1)].width = 25

# 保存
output_file = f'/tmp/zh/work/VisionSelector/qwen-evaluation/new_tasks_comparison_results.xlsx'
wb.save(output_file)
sys.stdout.write(f'Excel 已生成: {output_file}\n')
sys.stdout.flush()

# 打印结果摘要
sys.stdout.write('\n=== 结果汇总 ===\n')
for model_label, results in all_results.items():
    sys.stdout.write(f'\n[{model_label}]\n')
    for task in TASKS:
        metrics = metric_map.get(task, [])
        cell_value = format_cell(results, task, metrics)
        sys.stdout.write(f'  {task}: {cell_value}\n')
    sys.stdout.flush()
