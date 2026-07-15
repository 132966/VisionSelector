import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ROOT_DIR = '/tmp/zh/work/VisionSelector/qwen-evaluation/result/eval_time'
MODEL_NAME = 'VisionSelector-Qwen2.5-VL-3B-train-LayerMid-Attn-1epoch'

# 每个任务的指标列表（每个指标对应一列）
TASK_METRICS = [
    ('docvqa_val',    'anls,none',                    'DocVQA'),
    ('chartqa',       'relaxed_overall,none',         'ChartQA'),
    ('textvqa_val',   'exact_match,none',             'TextVQA'),
    ('ocrbench',      'ocrbench_accuracy,none',       'OCRBench'),
    ('scienceqa_img', 'exact_match,none',             'ScienceQA'),
    ('ai2d_no_mask',  'exact_match,flexible-extract', 'AI2D'),
    ('mmmu_val',      'mmmu_acc,none',                'MMMU'),
    ('mme',           'mme_cognition_score,none',     'MME-Cognition'),
    ('mme',           'mme_perception_score,none',    'MME-Perception'),
    ('mme',           None,                            'MME-Total'),  # cognition + perception
    ('pope',          'pope_accuracy,none',            'POPE-Accuracy'),
    ('pope',          'pope_precision,none',           'POPE-Precision'),
    ('pope',          'pope_recall,none',              'POPE-Recall'),
    ('pope',          'pope_f1_score,none',            'POPE-F1'),
    ('pope',          'pope_yes_ratio,none',           'POPE-YesRatio'),
]

TASKS = list(dict.fromkeys(t for t, _, _ in TASK_METRICS))  # 去重保序

# 收集结果
results = {}
for task in TASKS:
    result_path = f'{ROOT_DIR}/{task}/0.2/{MODEL_NAME}_selector_0.2_3b/output_ckpt__{MODEL_NAME}'
    if os.path.exists(result_path):
        json_files = [f for f in os.listdir(result_path) if f.endswith('_results.json')]
        if json_files:
            with open(os.path.join(result_path, json_files[0]), 'r') as f:
                data = json.load(f)
            task_results = data.get('results', {}).get(task, {})
            results[task] = task_results

# 创建 Excel
wb = Workbook()
ws = wb.active
ws.title = 'LayerMid-Attn-1epoch'

# 样式设置
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ['Dataset'] + [display for _, _, display in TASK_METRICS]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 数据行
row = 2
cell = ws.cell(row=row, column=1, value='LayerMid-Attn-1epoch')
cell.font = header_font
cell.alignment = center_align
cell.border = thin_border

for col, (task, metric, display) in enumerate(TASK_METRICS, 2):
    if task not in results:
        cell = ws.cell(row=row, column=col, value='-')
    elif metric is None and display == 'MME-Total':
        # MME Total = cognition + perception
        cog = results[task].get('mme_cognition_score,none', 0)
        per = results[task].get('mme_perception_score,none', 0)
        value = cog + per
        cell = ws.cell(row=row, column=col, value=round(value, 2))
    elif metric and metric in results[task]:
        value = results[task][metric]
        if isinstance(value, (int, float)) and 0 <= value <= 1:
            value = value * 100
        cell = ws.cell(row=row, column=col, value=round(value, 2))
    else:
        cell = ws.cell(row=row, column=col, value='-')
    cell.alignment = center_align
    cell.border = thin_border

# 调整列宽
ws.column_dimensions['A'].width = 25
for col in range(2, len(TASK_METRICS) + 2):
    col_letter = chr(ord('A') + col - 1)
    ws.column_dimensions[col_letter].width = 16

# 保存
output_file = f'/tmp/zh/work/VisionSelector/qwen-evaluation/{MODEL_NAME}_results.xlsx'
wb.save(output_file)

# 输出结果
sys.stdout.write(f'Excel 已生成: {output_file}\n')
sys.stdout.flush()
sys.stdout.write('\n=== LayerMid-Attn-1epoch 结果汇总 ===\n')
sys.stdout.flush()
for task, metric, display in TASK_METRICS:
    if task not in results:
        sys.stdout.write(f'{display}: -\n')
    elif metric is None and display == 'MME-Total':
        cog = results[task].get('mme_cognition_score,none', 0)
        per = results[task].get('mme_perception_score,none', 0)
        value = cog + per
        sys.stdout.write(f'{display}: {value:.2f}\n')
    elif metric and metric in results[task]:
        value = results[task][metric]
        if isinstance(value, (int, float)) and 0 <= value <= 1:
            value = value * 100
        sys.stdout.write(f'{display}: {value:.2f}\n')
    else:
        sys.stdout.write(f'{display}: -\n')
    sys.stdout.flush()
