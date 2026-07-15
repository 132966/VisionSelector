import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ROOT_DIR = '/tmp/zh/work/VisionSelector/qwen-evaluation/result/eval_time'
MODEL_NAME = 'VisionSelector-Qwen2.5-VL-3B-train-V2-bidirectional-10epoch'
TASKS = ['docvqa_val', 'chartqa', 'textvqa_val', 'ocrbench', 'scienceqa_img', 'ai2d_no_mask', 'mmmu_val', 'mme', 'pope']

# 指标映射
metric_map = {
    'docvqa_val': 'anls,none',
    'chartqa': 'relaxed_overall,none',
    'textvqa_val': 'exact_match,none',
    'ocrbench': 'ocrbench_accuracy,none',
    'scienceqa_img': 'exact_match,none',
    'ai2d_no_mask': 'accuracy,none',
    'mmmu_val': 'accuracy,none',
    'mme': 'score,none',
    'pope': 'score,none'
}

# 收集结果
results = {}
for task in TASKS:
    result_path = f'{ROOT_DIR}/{task}/0.2/{MODEL_NAME}_selector_0.2_3b/output_ckpt__{MODEL_NAME}'
    if os.path.exists(result_path):
        json_files = [f for f in os.listdir(result_path) if f.endswith('_results.json')]
        if json_files:
            with open(os.path.join(result_path, json_files[0]), 'r') as f:
                data = json.load(f)
            task_results = data.get('results', {}).get(task, {})
            results[task] = task_results

# 创建 Excel
wb = Workbook()
ws = wb.active
ws.title = MODEL_NAME[-20:]  # 截断长名称

# 样式设置
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ['Dataset'] + [t.replace('_val', '').replace('_', ' ').title() for t in TASKS]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 数据行
row = 2
cell = ws.cell(row=row, column=1, value='V2-bidirectional-10epoch')
cell.font = header_font
cell.alignment = center_align
cell.border = thin_border

for col, task in enumerate(TASKS, 2):
    if task in results and metric_map.get(task) in results[task]:
        value = results[task][metric_map[task]]
        # 0-1 范围的值乘以 100
        if isinstance(value, (int, float)) and 0 <= value <= 1:
            value = value * 100
        cell = ws.cell(row=row, column=col, value=round(value, 2))
    else:
        cell = ws.cell(row=row, column=col, value='-')
    cell.alignment = center_align
    cell.border = thin_border

# 调整列宽
ws.column_dimensions['A'].width = 25
for col in range(2, len(TASKS) + 2):
    col_letter = chr(ord('A') + col - 2) if col <= 26 else chr(ord('A') - 1 + col // 26) + chr(ord('A') + col % 26)
    ws.column_dimensions[col_letter].width = 14

# 保存
output_file = f'/tmp/zh/work/VisionSelector/qwen-evaluation/{MODEL_NAME}_results.xlsx'
wb.save(output_file)

# 输出结果
print(f'Excel 已生成: {output_file}')
print('\n=== V2-bidirectional-10epoch 结果汇总 ===')
for task in TASKS:
    if task in results and metric_map.get(task) in results[task]:
        value = results[task][metric_map[task]]
        if isinstance(value, (int, float)) and 0 <= value <= 1:
            value = value * 100
        print(f'{task}: {value:.2f}')
    else:
        print(f'{task}: -')
